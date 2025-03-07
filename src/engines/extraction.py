import pandas as pd
import fitz
import json
from datetime import datetime
import re
from pathlib import Path
import os
from openpyxl import load_workbook

from engines.prompt_catalog import *


class ExtractionEngine:
    def __init__(self, config, llm, embedding_model):
        self.config = config
        self.llm = llm
        self.embedding_model = embedding_model
        self.retrieve_masterdata()
    
    def read_document(self, filepath):
        doc = fitz.open(filepath)
        text = ""
        for page in doc:
            text += page.get_text()
        return text
    
    def save_extracted_data(self, df, output_path=None, sheetname="cv_data"):
        if output_path is None:
            output_path=self.config['data']['cv_data']

        if os.path.exists(output_path):

            cv_file = load_workbook(output_path, enumerate)
            sheet = cv_file[sheetname]
            row_count = sheet.max_row
            cv_file.close()

            with pd.ExcelWriter(output_path, mode='a', if_sheet_exists='overlay') as writer:  
                df.to_excel(writer,sheet_name=sheetname, startrow=row_count, index=False, header= False)
            
        else:
            df.to_excel(output_path, sheet_name=sheetname, index=False)
    
    @staticmethod
    def extract_json(text):
        json_pattern = r'\{(?:[^{}]|\{[^{}]*\})*\}'
        match = re.search(json_pattern, text)
        if match:
            return match.group(0)
        else:
            return "{}"
    
    def extract_jd(self, filepath):
        jd_text = self.read_document(filepath)
        user_prompt = user_prompt_jd_extraction.format(jd=jd_text)
        jd_data_str = self.llm.generate(user_prompt, system_prompt_jd_extraction)
        jd_data_str = self.extract_json(jd_data_str)
        jd_data_json = json.loads(jd_data_str)
        jd_data_json['filepath'] = filepath
        jd_data_json['years_of_experience'] = str(jd_data_json['years_of_experience'])
        refined_jd_str = self.llm.generate(user_prompt_jd_extraction.format(jd=jd_data_str), system_prompt_jd_summarize)
        refined_jd_str = self.extract_json(refined_jd_str)
        refined_jd_json = json.loads(refined_jd_str)
        print(refined_jd_json)
        if 'years_of_experience' in jd_data_json: 
            ksa_list = ["min " + str(jd_data_json['years_of_experience']) + " years of experience"]
            # del refined_jd_json['years_of_experience']
        for key in list(refined_jd_json.keys()):
            if isinstance(refined_jd_json[key], (list, tuple)):
                ksa_list.extend(refined_jd_json[key])
            else:
                ksa_list.append(refined_jd_json[key])
        jd_data_json['ksa'] = list(set(ksa_list))
        return jd_data_json
    
    def extract_cv_from_folder(self, folder_path="../data/CV"):
        files_list = os.listdir(folder_path)
        for f in files_list:
            print(f)
            cv_data_json = self.extract_cv(os.path.join(folder_path, f))
            cv_df = pd.DataFrame([cv_data_json])
            self.save_extracted_data(cv_df)
    
    def extract_cv(self, filepath):
        cv_text = self.read_document(filepath)
        user_prompt = user_prompt_cv_extraction.format(cv=cv_text)
        cv_data_str = self.llm.generate(user_prompt, system_prompt_cv_extraction)
        cv_data_str = self.extract_json(cv_data_str)
        cv_data_json = json.loads(cv_data_str)
        embedding = self.embed_cv(cv_data_json)
        filename = Path(filepath).name
        employee_id = filename[:6]
        last_hire_date = self.retrieve_last_hire_date(employee_id)
        experience_years = self.extract_experience(cv_data_json, last_hire_date)
        cv_data_json['years_of_experience'] = str(experience_years)
        cv_data_json['last_hire_date'] = last_hire_date
        cv_data_json['employee_id'] = employee_id
        cv_data_json['filepath'] = filepath
        cv_data_json['embedding'] = embedding
        return cv_data_json
    
    def extract_experience(self, cv, last_hire_date):
        now = datetime.now().strftime("%d %b %Y")
        job_history = cv['job_history']
        system_prompt = system_prompt_cv_experience.format(last_hire_date=last_hire_date, now=now)
        user_prompt = user_prompt_cv_experience.format(last_hire_date=last_hire_date, job_history=job_history)
        experience_str = self.llm.generate(user_prompt, system_prompt)
        pattern = r'\{[^{}]*"years_of_experience"[^{}]*\}'
        match = re.findall(pattern, experience_str)
        if match:
            years_of_experience_str = match[-1]
            years_of_experience_json = json.loads(years_of_experience_str)
            years_of_experience = years_of_experience_json['years_of_experience']
        else:
            years_of_experience = None
        return years_of_experience
    
    def retrieve_masterdata(self):
        self.employees_df = pd.read_excel(self.config['data']['employees_master_data'])
        self.employees_df['Serial Number'] = self.employees_df['Serial Number'].astype(str)
        
    def retrieve_last_hire_date(self, employee_id):
        last_hire_date = self.employees_df.loc[self.employees_df['Serial Number']==str(employee_id), "Last Hire Date"].dt.strftime("%d %b %Y").values[0]
        return last_hire_date
    
    def embed_cv(self, cv_data_json):
        cv_data_json = cv_data_json.copy()
        user_prompt = user_prompt_cv_summarize.format(cv=str(cv_data_json))
        system_prompt = system_prompt_cv_summarize
        cv_summary = self.llm.generate(user_prompt, system_prompt)
        cv_embedding = self.embedding_model.embed(cv_summary)
        return cv_embedding